"""Parse documents into a consistent sequence of blocks.

Every parser returns the same Block structure so chunking remains one shared
implementation. Format differences are represented by block kind and metadata.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

# ------------------------------------------------------------------ Model
Kind = str  # "heading" | "para" | "row" | "question" | "answer"


@dataclass
class Block:
    kind: Kind
    text: str
    level: int | None = None          # Heading level: 1-3.
    page: int | None = None
    sheet: str | None = None
    row: int | None = None
    meta: dict = field(default_factory=dict)


class UnsupportedFormat(ValueError):
    pass


# ------------------------------------------------------------------ PDF
_HEB_HEADING_NUM = re.compile(r"^\s*\d+(\.\d+)*\s")


def parse_pdf(path: Path) -> list[Block]:
    """Detect PDF headings from font size and bold styling.

    PDFs contain positioned text rather than semantic structure. The most common
    font size is body text; significantly larger or bold numbered text is treated as a heading.
    """
    import pymupdf

    doc = pymupdf.open(path)
    raw: list[tuple[str, float, bool, int]] = []  # text, size, bold, page

    for page_no, page in enumerate(doc, start=1):
        for block in page.get_text("dict")["blocks"]:
            if block.get("type") != 0:  # 0 = text.
                continue
            for line in block["lines"]:
                spans = [s for s in line["spans"] if s["text"].strip()]
                if not spans:
                    continue
                text_ = "".join(s["text"] for s in spans).strip()
                size = max(s["size"] for s in spans)
                # ביט 4 בדגלי pymupdf מסמן bold
                bold = any(bool(s["flags"] & 2**4) for s in spans)
                raw.append((text_, round(size, 1), bold, page_no))
    doc.close()

    if not raw:
        return []

    # Body size is the size covering the most characters, not the most lines.
    weights: Counter[float] = Counter()
    for text_, size, _bold, _pg in raw:
        weights[size] += len(text_)
    body_size = weights.most_common(1)[0][0]

    heading_sizes = sorted({s for _t, s, _b, _p in raw if s > body_size * 1.12}, reverse=True)

    blocks: list[Block] = []
    for text_, size, bold, page_no in raw:
        level: int | None = None
        if size in heading_sizes:
            level = min(heading_sizes.index(size) + 1, 3)
        elif bold and size >= body_size and _HEB_HEADING_NUM.match(text_) and len(text_) < 90:
            level = 3
        if level:
            blocks.append(Block("heading", text_, level=level, page=page_no))
        else:
            blocks.append(Block("para", text_, page=page_no))
    return blocks


# ------------------------------------------------------------------ DOCX
def parse_docx(path: Path) -> list[Block]:
    """Word provides real semantic structure through Heading styles."""
    import docx
    from docx.document import Document as DocxDocument
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document: DocxDocument = docx.Document(str(path))
    blocks: list[Block] = []

    def iter_body(parent):
        """Yield paragraphs and tables in document order."""
        body = parent.element.body
        for child in body.iterchildren():
            if child.tag == qn("w:p"):
                yield Paragraph(child, parent)
            elif child.tag == qn("w:tbl"):
                yield Table(child, parent)

    for item in iter_body(document):
        if isinstance(item, Paragraph):
            text_ = item.text.strip()
            if not text_:
                continue
            # Detect question and answer labels before styles: in FAQs a
            # question may look like a heading but is half of an information unit.
            if text_.startswith("שאלה:"):
                blocks.append(Block("question", text_.removeprefix("שאלה:").strip()))
                continue
            if text_.startswith("תשובה:"):
                blocks.append(Block("answer", text_.removeprefix("תשובה:").strip()))
                continue

            # A paragraph may have no defined style.
            style = (getattr(item.style, "name", None) or "").lower()
            if style.startswith("heading"):
                digits = "".join(ch for ch in style if ch.isdigit())
                level = min(int(digits or 1), 3)
                blocks.append(Block("heading", text_, level=level))
            else:
                blocks.append(Block("para", text_))
        else:  # Table
            rows = [[c.text.strip() for c in r.cells] for r in item.rows]
            if not rows:
                continue
            header, *body_rows = rows
            for i, row in enumerate(body_rows, start=1):
                pairs = [f"{h}: {v}" for h, v in zip(header, row, strict=False) if v]
                if pairs:
                    blocks.append(Block("row", " | ".join(pairs), row=i, meta={"from": "table"}))
    return blocks


# ------------------------------------------------------------------ XLSX
def parse_xlsx(path: Path) -> list[Block]:
    """Each spreadsheet row is an independent information unit and block.

    Column headers are included in each row; otherwise values such as
    "7 | 18500 | 24300" lack meaning to both people and models.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    blocks: list[Block] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        blocks.append(Block("heading", f"גיליון: {ws.title}", level=1, sheet=ws.title))

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        for idx, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            pairs = [
                f"{h}: {v}"
                for h, v in zip(header, row, strict=False)
                if v is not None and str(v).strip() and h
            ]
            if pairs:
                blocks.append(
                    Block("row", " | ".join(pairs), sheet=ws.title, row=idx)
                )
    wb.close()
    return blocks


# ------------------------------------------------------------------ HTML
def parse_html(path: Path) -> list[Block]:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "lxml")
    blocks: list[Block] = []

    for el in soup.find_all(["h1", "h2", "h3", "h4", "p", "li", "tr"]):
        text_ = el.get_text(" ", strip=True)
        if not text_:
            continue
        name = el.name
        classes = el.get("class") or []
        # As in Word, a question may be styled as a heading without being a section boundary.
        if "q" in classes:
            blocks.append(Block("question", text_))
        elif "a" in classes:
            blocks.append(Block("answer", text_))
        elif name in {"h1", "h2", "h3", "h4"}:
            blocks.append(Block("heading", text_, level=min(int(name[1]), 3)))
        elif name == "tr":
            cells = [c.get_text(" ", strip=True) for c in el.find_all(["td", "th"])]
            if any(cells):
                blocks.append(Block("row", " | ".join(c for c in cells if c)))
        else:
            blocks.append(Block("para", text_))
    return blocks


# ------------------------------------------------------------------ Dispatcher
PARSERS = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".html": parse_html,
    ".htm": parse_html,
}


def parse(path: Path) -> list[Block]:
    parser = PARSERS.get(path.suffix.lower())
    if parser is None:
        raise UnsupportedFormat(f"No parser for extension {path.suffix}")
    return parser(path)
